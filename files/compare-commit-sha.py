"""
This script processes and compares two Excel files containing branch and commit information. It performs the following tasks:

1. Loads two Excel files ('Branches_and_Commit_Source.xlsx' and 'Branches_and_Commit_Target.xlsx') into pandas DataFrames.
2. Merges these DataFrames to identify and classify differences between the source and target files.
3. Adds 'colour-code' and 'Remark' columns based on the comparison:
   - 'RED' for entries present only in the source file.
   - 'GREEN' for 'dependabot/' branches present only in the target file.
   - 'YELLOW' for other branches present only in the target file.
4. Drops the merge indicator column and saves the updated DataFrame to a temporary Excel file.
5. Appends legend information explaining the color codes to the worksheet.
6. Applies cell coloring based on 'colour-code' values and centers the legend text.
7. Saves the final annotated workbook as 'final_report.xlsx'.
8. Removes the temporary Excel file used during processing.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Load the two Excel files
df_source = pd.read_excel('Branches_and_Commit_Source.xlsx')
df_target = pd.read_excel('Branches_and_Commit_Target.xlsx')

# Merge the two dataframes, indicator=True will add a column '_merge' that indicates the source of each row
df = pd.merge(df_source, df_target, on=['Repository', 'Branch', 'Commit SHA'], how='outer', indicator=True)

# Create 'colour-code' and 'Remark' columns based on the conditions
df['colour-code'] = 'white'
df['Remark'] = 'Good'

df.loc[df['_merge'] == 'left_only', 'colour-code'] = 'RED'
df.loc[df['_merge'] == 'left_only', 'Remark'] = 'FAILED'

df.loc[(df['_merge'] == 'right_only') & df['Branch'].str.startswith('dependabot/'), 'colour-code'] = 'GREEN'
df.loc[(df['_merge'] == 'right_only') & df['Branch'].str.startswith('dependabot/'), 'Remark'] = 'SUCCESSFUL'

df.loc[(df['_merge'] == 'right_only') & ~df['Branch'].str.startswith('dependabot/'), 'colour-code'] = 'YELLOW'
df.loc[(df['_merge'] == 'right_only') & ~df['Branch'].str.startswith('dependabot/'), 'Remark'] = 'MANUAL REVIEW'

# Drop the '_merge' column
df = df.drop(columns='_merge')

# Save the main DataFrame to a temporary Excel file
temp_file = 'report.xlsx'
df.to_excel(temp_file, index=False)

# Load the workbook and the first worksheet
wb = load_workbook(temp_file)
ws = wb.active

# Define the legends as a DataFrame
legends = pd.DataFrame({
    'Repository': [''] * 3,
    'Branch': [''] * 3,
    'Commit SHA': [''] * 3,
    'colour-code': ['GREEN', 'RED', 'YELLOW'],
    'Remark': [
        'Entries which are added at target but not present in Source. These entries are Dependabot related',
        'Entries which are present in Source but not present in Target',
        'Entries which are added at target but not present in Source. These entries are not Dependabot related'
    ]
})

# Append a blank row
ws.append([''] * len(df.columns))

# Append the legends to the worksheet
for r in dataframe_to_rows(legends, index=False, header=False):
    ws.append(r)

# Define colors for the color codes
color_map = {
    'GREEN': '00FF00',
    'RED': 'FF0000',
    'YELLOW': 'FFFF00',
}

# Apply cell coloring based on 'colour-code' values in the main data
for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=4, max_col=4):
    for cell in row:
        fill_color = color_map.get(cell.value, 'FFFFFF')
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# Apply cell coloring to the legend
legend_start_row = len(df) + 3
for row in ws.iter_rows(min_row=legend_start_row, max_row=legend_start_row+2, min_col=4, max_col=4):
    for cell in row:
        fill_color = color_map.get(cell.value, 'FFFFFF')
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# Center align the legend text
for row in ws.iter_rows(min_row=legend_start_row, max_row=legend_start_row+2, min_col=5, max_col=5):
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# Save the final workbook
final_file = 'final_report.xlsx'
wb.save(final_file)

# Remove the temporary file
os.remove(temp_file)