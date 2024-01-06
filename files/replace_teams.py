import openpyxl
import os
from github import Github

# Read repository information from the first Excel file
workbook = openpyxl.load_workbook('data/CodeOwners-File-Location.xlsx')
sheet = workbook.active

codeowners_file_location_path = 'data/CodeOwners-File-Location.xlsx'
teams_to_replace_path = 'data/CodeOwners-Team-Replacement.xlsx'

for row in sheet.iter_rows(min_row=2, values_only=True):
    org_name, repo_name, codeowners_path = row

    # Clone the repository
    repo_path = f'temp_repos/{repo_name}'
    if not os.path.exists("temp_repos"):
        os.makedirs(repo_path, exist_ok=True)

    os.system(f'git clone https://github.com/{org_name}/{repo_name}.git {repo_path}')
    os.chdir(repo_path)

    # Read teams to be replaced from the second Excel file
    with openpyxl.load_workbook(teams_to_replace_path) as teams_workbook:
        teams_sheet = teams_workbook.active
        for team_row in teams_sheet.iter_rows(min_row=2, values_only=True):
            search_team, replace_team = team_row

            # Replace team names in CODEOWNERS file
            codeowners_file = os.path.join(repo_path, codeowners_path)
            with open(codeowners_file, 'r') as f:
                codeowners_content = f.read()
            codeowners_content = codeowners_content.replace(search_team, replace_team)
            with open(codeowners_file, 'w') as f:
                f.write(codeowners_content)

            # Commit and push changes
            os.system('git add .')
            os.system(f'git commit -m "Replace {search_team} with {replace_team}"')
            os.system('git push')

            print(f'Teams replaced in {org_name}/{repo_name}')

    os.chdir('repo/central_repo')

workbook.close()
